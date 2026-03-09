import time
import os

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from typing import Dict, List, Optional

from PyQt6.QtWidgets import *
from PyQt6.QtCore import *
from PyQt6.QtGui import *

from api.owencloud import OwenCloudClient
from models.data_model import DataModel
from models.events import EventEngine, EventStorage
from config_manager import ConfigManager
from utils.paths import resource_path
from constants import EVENT_LABELS
from datetime import datetime, timedelta
from ui.dashboard import DashboardPage
from ui.events_page import EventsPage
from ui.settings_page import SettingsPage
from ui.login_dialog import LoginDialog
from ui.map_page import MapPage
from ui.theme import ThemeManager


class MainWindow(QMainWindow):
    """Главное SCADA-окно."""

    def on_settings_changed(self):
        print("Настройки изменены → применяем новое сопоставление кодов.")
        self.apply_config_overrides()

    def __init__(
        self,
        api_client: OwenCloudClient,
        device_id: int,
        param_ids_by_key: Dict[str, int],
        devices: List[Dict],
    ):
        super().__init__()
        self.theme_mgr = ThemeManager()
        self.setWindowIcon(QIcon(resource_path("assets/icon.ico")))
        self.devices = devices
        self.api_client = api_client
        self.device_id = device_id
        self.param_ids_by_key = param_ids_by_key
        self.config_manager = ConfigManager(device_id)  # менеджер конфигурации для этого устройства

        self.apply_config_overrides()  # применяем переопределения из файла настроек (если есть)

        self.setWindowTitle("Термовизор — Панель управления")
        self.resize(1400, 750)
        self._history_loaded = False
        self.data_model = DataModel()
        self.event_storage = EventStorage(self.device_id)
        self.event_engine = EventEngine(self.event_storage, max_events=200)

        central = QWidget()
        main_layout = QHBoxLayout(central)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # Боковое меню
        self.menu_frame = QFrame()
        self.menu_frame.setFixedWidth(220)
        self.menu_frame.setObjectName("SideMenu")
        self.menu_frame.setStyleSheet("")
        menu_layout = QVBoxLayout(self.menu_frame)
        menu_layout.setContentsMargins(10, 20, 10, 20)
        menu_layout.setSpacing(10)

        logo_lbl = QLabel("Термовизор")
        logo_lbl.setFont(QFont("Arial", 14, QFont.Weight.Bold))
        logo_lbl.setStyleSheet("")
        menu_layout.addWidget(logo_lbl)
        menu_layout.addSpacing(20)

        self.buttons: List[QPushButton] = []

        # Верхние пункты (без "Настройки")
        top_items = [
            ("Главная", "🏠", 0),
            ("События", "📝", 1),
            ("Карта", "🗺️", 3),
        ]

        for text, icon, page_index in top_items:
            btn = QPushButton(f"{icon}  {text}")
            btn.setObjectName("NavButton")
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.setCheckable(True)
            btn.setStyleSheet("")
            btn.clicked.connect(lambda checked, i=page_index: self.set_page(i))
            self.buttons.append(btn)
            menu_layout.addWidget(btn)

        # Проталкиваем вниз
        menu_layout.addStretch()

        # Нижняя отдельная кнопка "Настройки"
        settings_btn = QPushButton("⚙️  Настройки")
        settings_btn.setObjectName("NavButton")
        settings_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        settings_btn.setCheckable(True)
        settings_btn.setStyleSheet("")
        settings_btn.clicked.connect(lambda checked, i=2: self.set_page(i))
        self.buttons.insert(2, settings_btn)  # чтобы индекс 2 соответствовал странице настроек
        menu_layout.addWidget(settings_btn)
        main_layout.addWidget(self.menu_frame)

        # Страницы
        self.stack = QStackedWidget()

        # 0 — Главная
        self.dashboard_page = DashboardPage()
        self.stack.addWidget(self.dashboard_page)
        self.dashboard_page.set_devices(self.devices, self.device_id)
        self.dashboard_page.device_changed.connect(self.on_device_changed)
        self.dashboard_page.logout_requested.connect(self.on_logout_clicked)
        self.dashboard_page.boiler_changed.connect(self.on_boiler_changed)
        self.dashboard_page.range_changed.connect(self.on_range_changed)
        self.dashboard_page.period_changed.connect(self.on_period_changed)
        self.dashboard_page.live_clicked.connect(self.on_live_clicked)
        self.dashboard_page.theme_btn.clicked.connect(self.on_theme_toggle_clicked)
        self.dashboard_page.apply_theme(self.theme_mgr.get_theme())
        self.dashboard_page.archive_requested.connect(self.on_archive_requested)

        # 1 — События
        self.events_page = EventsPage()
        self.stack.addWidget(self.events_page)

        # 2 — Настройки
        self.settings_page = SettingsPage(self.api_client, self.device_id, self.config_manager)
        self.settings_page.settings_saved.connect(self.on_settings_changed)
        self.stack.addWidget(self.settings_page)

        # 3 — Карта
        self.map_page = MapPage()
        self.stack.addWidget(self.map_page)


        main_layout.addWidget(self.stack, 1)
        self.setCentralWidget(central)

        self.buttons[0].setChecked(True)
        self.stack.setCurrentIndex(0)

        # Таймер опроса OwenCloud
        self.owen_timer = QTimer(self)
        self.owen_timer.setInterval(10_000) # каждые 10 секунд
        self.owen_timer.timeout.connect(self.update_from_owencloud)
        self.owen_timer.start()
        QTimer.singleShot(0, self.update_from_owencloud)
        self._history_loaded_for: Optional[int] = None

    def on_archive_requested(self, boiler_n: int, start_ts: int, end_ts: int):
        # какие PID нам нужны (как для графика)
        pid_supply = self.param_ids_by_key.get(f"boiler_{boiler_n}")
        pid_return = self.param_ids_by_key.get(f"boiler_{boiler_n}_return")
        pid_press = self.param_ids_by_key.get(f"boiler_{boiler_n}_pressure")

        ids = [pid for pid in (pid_supply, pid_return, pid_press) if pid]

        if not ids:
            QMessageBox.warning(self, "Ошибка", "Нет настроенных параметров для выбранного котла (PID).")
            return

        # выбираем куда сохранить файл
        default_name = f"archive_boiler_{boiler_n}_{datetime.fromtimestamp(start_ts).strftime('%Y%m%d_%H%M')}-{datetime.fromtimestamp(end_ts).strftime('%Y%m%d_%H%M')}.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Сохранить архив в Excel",
            os.path.join(os.path.expanduser("~"), "Downloads", default_name),
            "Excel (*.xlsx)"
        )
        if not path:
            return

        # грузим историю шагом 60 сек
        try:
            hist = self.api_client.get_history(
                ids,
                start_ts,
                end_ts,
                step_sec=60,
            )
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось получить историю: {e}")
            return

        series_supply = hist.get(pid_supply, []) if pid_supply else []
        series_return = hist.get(pid_return, []) if pid_return else []
        series_press = hist.get(pid_press, []) if pid_press else []

        # превращаем в dict minute_ts -> value (на всякий случай нормализуем к минуте)
        def to_minute_dict(series: list[tuple[float, float]]) -> dict[int, float]:
            d: dict[int, float] = {}
            for ts, v in series:
                try:
                    mt = (int(ts) // 60) * 60
                    d[mt] = float(v)
                except Exception:
                    continue
            return d

        d_sup = to_minute_dict(series_supply)
        d_ret = to_minute_dict(series_return)
        d_pre = to_minute_dict(series_press)

        # строим минутную сетку строго от "С" до "По" (включительно)
        # (в UI формат dd.MM.yyyy HH:mm, секунды обычно = 0, но всё равно нормализуем)
        start_ts = (int(start_ts) // 60) * 60
        end_ts = (int(end_ts) // 60) * 60

        # создаём Excel
        wb = Workbook()
        ws = wb.active
        ws.title = f"Котёл {boiler_n}"

        headers = ["Дата и время", "Температура подачи", "Температура обратки", "Давление подачи"]
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = Font(bold=True)

        last_sup = 0.0
        last_ret = 0.0
        last_pre = 0.0

        cur = start_ts
        while cur <= end_ts:
            if cur in d_sup:
                last_sup = d_sup[cur]
            if cur in d_ret:
                last_ret = d_ret[cur]
            if cur in d_pre:
                last_pre = d_pre[cur]

            dt = datetime.fromtimestamp(cur)
            ws.append([dt.strftime("%d.%m.%Y %H:%M"), last_sup, last_ret, last_pre])
            cur += 60

        # чуть удобнее: ширины колонок
        widths = [20, 20, 20, 16]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        try:
            wb.save(path)
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить файл:\n{e}")
            return

        QMessageBox.information(self, "Готово", f"Архив выгружен:\n{path}")

    def on_theme_toggle_clicked(self):
        new_theme = self.theme_mgr.toggle()

        # применяем QSS ко всему приложению
        app = QApplication.instance()
        if app:
            self.theme_mgr.apply_to_app(app)

        # обновляем то, что не покрывается QSS (pyqtgraph)
        self.dashboard_page.apply_theme(new_theme)

        # можно менять значок кнопки
        if hasattr(self.dashboard_page, "theme_btn"):
            self.dashboard_page.theme_btn.setText("🌙" if new_theme == "dark" else "☀")

    def on_period_changed(self, start_ts: int, end_ts: int):
        dp = self.dashboard_page
        dp.period_mode = True
        dp.period_start_ts = int(start_ts)
        dp.period_end_ts = int(end_ts)

        self._history_loaded = False
        self._history_loaded_for = None

        self.load_initial_history()
        self.update_graph()

    def on_live_clicked(self):
        dp = self.dashboard_page
        dp.period_mode = False
        dp.period_start_ts = None
        dp.period_end_ts = None

        self._history_loaded = False
        self._history_loaded_for = None

        self.load_initial_history()  # загрузит последние N часов (Диапазон)
        self.update_graph()  # live-режим продолжит добавлять точки

    def on_device_changed(self, new_device_id: int):
        if new_device_id == self.device_id:
            return

        print(f"🔁 Переключение на устройство {new_device_id}")

        # останавливаем опрос
        self.owen_timer.stop()

        # меняем устройство
        self.device_id = new_device_id
        self.config_manager = ConfigManager(new_device_id)

        # сбрасываем модель
        self.data_model = DataModel()

        # новая БД событий
        self.event_storage = EventStorage(new_device_id)
        self.event_engine = EventEngine(self.event_storage, max_events=200)

        # перечитываем настройки
        self.apply_config_overrides()
        # сбрасываем историю при смене устройства
        self._history_loaded = False
        self._history_loaded_for = None

        # обновляем GUI
        self.update_gui_from_model()
        self.events_page.update_local_events(self.event_engine.events)
        # запускаем опрос заново
        self.owen_timer.start()

    def on_boiler_changed(self, n: int):
        # разрешаем загрузку истории заново
        self._history_loaded = False
        self._history_loaded_for = None

        # грузим историю и сразу рисуем
        self.load_initial_history()
        self.update_graph()

    def on_range_changed(self, hours: int):
        # разрешаем загрузку истории заново
        self._history_loaded = False
        self._history_loaded_for = None

        # грузим историю и сразу рисуем
        self.load_initial_history()
        self.update_graph()

    def on_logout_clicked(self):
        msg = QMessageBox(self)
        msg.setWindowTitle("Выход из аккаунта")
        msg.setText("Вы действительно хотите выйти из аккаунта?")
        msg.setIcon(QMessageBox.Icon.Question)

        msg.setStandardButtons(
            QMessageBox.StandardButton.Yes |
            QMessageBox.StandardButton.No
        )
        msg.setDefaultButton(QMessageBox.StandardButton.No)
        msg.button(QMessageBox.StandardButton.Yes).setText("Да")
        msg.button(QMessageBox.StandardButton.No).setText("Нет")

        # 🎨 ТЁМНАЯ ТЕМА + БЕЛЫЙ ТЕКСТ
        msg.setStyleSheet("""
            QMessageBox {
                background-color: #101215;
            }
            QLabel {
                color: white;
                font-size: 11pt;
            }
            QPushButton {
                background-color: #1F2226;
                color: white;
                border-radius: 6px;
                padding: 6px 14px;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #E53935;
            }
        """)

        reply = msg.exec()

        if reply != QMessageBox.StandardButton.Yes:
            return
        self.owen_timer.stop()
        self.close()
        self.open_login_again()

    def open_login_again(self):
        login = LoginDialog()
        login.showMaximized()

        if login.exec() != QDialog.DialogCode.Accepted:
            # если пользователь закрыл окно входа
            QApplication.quit()
            return

        # создаём НОВОЕ главное окно
        win = MainWindow(
            login.api_client,
            login.selected_device_id,
            login.param_ids_by_key,
            login.devices
        )
        win.showMaximized()

    # ---------- Навигация ----------

    def set_page(self, index: int):
        for i, btn in enumerate(self.buttons):
            btn.setChecked(i == index)
        self.stack.setCurrentIndex(index)

    # ---------- Утилиты ----------

    @staticmethod
    def status_color(status: str) -> str:
        return {
            "Работа": "#4CAF50",
            "Остановка": "#757575",
            "Авария": "#E53935",
        }.get(status, "#9E9E9E")

    # ---------- Обновление GUI из модели ----------

    def update_status_card(self, key: str, status: str):
        card = self.dashboard_page.card_refs.get(key)
        if not card:
            return
        card.update_card(
            status=status,
            color=self.status_color(status),
        )

    def update_gui_from_model(self):
        m = self.data_model
        cards = self.dashboard_page.card_refs

        # ============================
        #   КОТЛОВОЙ КОНТУР
        # ============================
        for i in range(1, 3):
            pump = m.kotlovoy[f"pump{i}"]
            self.update_status_card(f"kotlovoy_pump{i}", pump.status)

        # ============================
        #   СЕТЕВОЙ КОНТУР
        # ============================
        for i in (1, 2):
            pump = m.network.__dict__[f"pump{i}"]
            self.update_status_card(f"network_pump{i}", pump.status)

        cards["network_pressure_before"].update_card(
            value=m.network.pressure_before,
            unit="Bar",
        )
        cards["network_pressure_after"].update_card(
            value=m.network.pressure_after,
            unit="Bar",
        )

        cards["network_temp_supply"].update_card(
            value=m.network.temp_supply,
            unit="°C",
        )
        cards["network_temp_return"].update_card(
            value=m.network.temp_return,
            unit="°C",
        )

        # ============================
        #   КОТЛЫ
        # ============================
        for i in range(1, 4):
            b = m.boilers[i]

            card = cards[f"boiler_{i}"]
            card.update_card(
                title=f"Котёл {i}",
                value=None,
                status=b.status,
                color=self.status_color(b.status),
            )

            temp = f"{b.temperature:.1f} °C"
            press = f"{b.pressure_return:.2f} Bar"
            card.value_label.setText(f"{temp} / {press}")
            card.value_label.show()

            cards[f"boiler_{i}_bunker"].set_value(b.bunker_level)

        # ============================
        #   ДЫМОСОСЫ
        # ============================
        for i in (1, 2):
            d = m.dymososy[i]
            self.update_status_card(f"dymosos_{i}", d.status)

        # ============================
        #   ТЕМПЕРАТУРА УЛИЦЫ
        # ============================
        self.dashboard_page.outdoor_temp_label.setText(
            f"🌤️ Улица: {int(m.outdoor_temperature)} °C"
        )

    # ---------- Получение данных с OwenCloud и обновление модели ----------

    def apply_config_overrides(self):
        """
        Читает файл configs/device_<ID>.json и полностью формирует param_ids_by_key.
        Работает ТОЛЬКО с пользовательскими кодами. Если настроек нет — все значения будут 0.
        """
        config = self.config_manager.load()
        mapping = config.get("mapping")

        if not mapping:
            print("Настройки пустые → отключаем все параметры.")
            self.param_ids_by_key = {}
            return

        try:
            params = self.api_client.get_device_parameters(self.device_id)
        except Exception as e:
            print("Ошибка загрузки параметров:", e)
            self.param_ids_by_key = {}
            return

        # code → id
        code_to_id = {}
        for p in params:
            pid = p.get("id")
            code = p.get("code") or p.get("name")
            if pid and code:
                code_to_id[str(code)] = int(pid)

        new_map = {}

        for key, code in mapping.items():
            code = str(code).strip()
            pid = code_to_id.get(code)
            if pid is None:
                print(f"[WARN] Не найден параметр '{code}' для '{key}'")
                continue
            new_map[key] = pid

        self.param_ids_by_key = new_map

        print("\n=== КАРТА ПАРАМЕТРОВ (ПОСЛЕ НАСТРОЕК) ===")
        if not new_map:
            print("Нет корректных параметров → все данные будут 0")
        for k, v in new_map.items():
            print(f"{k} → {v}")

    @staticmethod
    def bool_to_status(v: float) -> str:
        return "Работа" if int(v) == 1 else "Остановка"

    @staticmethod
    def apply_alarm(status: str, alarm: Optional[float]) -> str:
        if alarm is not None and int(alarm) == 1:
            return "Авария"
        return status

    def update_from_owencloud(self):
        """
        Опрос OwenCloud:
            - получает значения нужных параметров
            - обновляет модель
            - обновляет GUI
        """

        # Если нет настроенных параметров → ничего не обновляем, показываем только нули
        if not self.param_ids_by_key:
            self.update_gui_from_model()
            return

        ids = list(self.param_ids_by_key.values())
        values_by_id = self.api_client.get_last_values_by_ids(ids)

        if not values_by_id:
            self.update_gui_from_model()
            return

        # ===============================
        #   ЛОКАЛЬНЫЕ СОБЫТИЯ (по изменениям)
        # ===============================
        for key, pid in self.param_ids_by_key.items():
            if pid not in values_by_id:
                continue

            val = values_by_id[pid]
            label = EVENT_LABELS.get(key)

            if not label:
                continue

            if key.endswith("_status") or key in ("dymosos_1", "dymosos_2"):
                self.event_engine.push(key=key, value=val, label=label, kind="status")

            if key.endswith("_alarm"):
                self.event_engine.push(key=key, value=val, label=label, kind="alarm")

        m = self.data_model

        def get_val(key: str) -> Optional[float]:
            pid = self.param_ids_by_key.get(key)
            if pid is None:
                return None
            return values_by_id.get(pid)

        for i in range(1, 4):
            v = get_val(f"boiler_{i}_bunker")
            if v is not None:
                m.boilers[i].bunker_level = max(0, min(100, float(v)))

        # ===============================
        #   ТЕМПЕРАТУРА НА УЛИЦЕ
        # ===============================
        v_outdoor = get_val("outdoor_temperature")
        if v_outdoor is not None:
            self.data_model.outdoor_temperature = v_outdoor
        # ===============================
        #   КОТЛОВОЙ КОНТУР
        # ===============================
        for i in range(1, 3):  # только насос 1..2
            status_val = get_val(f"kotlovoy_pump{i}_status")
            alarm_val = get_val(f"kotlovoy_pump{i}_alarm")

            if status_val is not None:
                status = self.bool_to_status(status_val)
                status = self.apply_alarm(status, alarm_val)
                m.kotlovoy[f"pump{i}"].status = status

        # ===============================
        #   СЕТЕВОЙ КОНТУР
        # ===============================
        v_network_before = get_val("network_pressure_before")
        if v_network_before is not None:
            m.network.pressure_before = v_network_before

        v_network_after = get_val("network_pressure_after")
        if v_network_after is not None:
            m.network.pressure_after = v_network_after

        t_network_supply = get_val("network_temp_supply")
        if t_network_supply is not None:
            m.network.temp_supply = t_network_supply

        t_network_return = get_val("network_temp_return")
        if t_network_return is not None:
            m.network.temp_return = t_network_return

        for i in (1, 2):
            pump = m.network.__dict__[f"pump{i}"]

            status_val = get_val(f"network_pump{i}_status")
            alarm_val = get_val(f"network_pump{i}_alarm")

            if status_val is not None:
                status = self.bool_to_status(status_val)
                status = self.apply_alarm(status, alarm_val)
                pump.status = status

        # ===============================
        #   КОТЛЫ
        # ===============================

        for i in range(1, 4):
            v = get_val(f"boiler_{i}_pressure")
            if v is not None:
                m.boilers[i].pressure_return = v
                m.boilers[i].pressure_supply = v

        for i in range(1, 4):
            b = m.boilers[i]

            temp_supply = get_val(f"boiler_{i}")
            temp_return = get_val(f"boiler_{i}_return")
            press_return = get_val(f"boiler_{i}_pressure")
            status_val = get_val(f"boiler_{i}_status")
            alarm_val = get_val(f"boiler_{i}_alarm")

            if temp_supply is not None:
                b.temperature = temp_supply
            if temp_return is not None:
                b.temp_return = temp_return
            if press_return is not None:
                b.pressure_return = press_return
                b.pressure_supply = press_return
            if status_val is not None:
                status = self.bool_to_status(status_val)
                status = self.apply_alarm(status, alarm_val)
                b.status = status

        # ===============================
        #   ДЫМОСОСЫ
        # ===============================
        for i in (1, 2):
            status_val = get_val(f"dymosos_{i}")
            alarm_val = get_val(f"dymosos_{i}_alarm")

            if status_val is not None:
                status = self.bool_to_status(status_val)
                status = self.apply_alarm(status, alarm_val)
                m.dymososy[i].status = status

        # Обновляем GUI

        self.update_gui_from_model()
        self.load_initial_history()
        self.update_graph()
        self.events_page.update_local_events(self.event_engine.events)

    def update_graph(self):
        dp = self.dashboard_page
        if getattr(dp, "period_mode", False):
            return
        m = self.data_model

        boiler_n = getattr(dp, "selected_boiler", 1)
        b = m.boilers.get(int(boiler_n))
        if not b:
            return

        t = time.time()
        WINDOW_SEC = int(getattr(dp, "graph_window_hours", 4)) * 60 * 60  # 1 - 24 часа

        # добавляем новую точку
        dp.graph_time.append(t)
        dp.graph_temp_supply.append(b.temperature)
        dp.graph_temp_return.append(b.temp_return)
        dp.graph_press_supply.append(b.pressure_return)

        # подрезаем всё, что старше 4 часов
        cut = t - WINDOW_SEC
        while dp.graph_time and dp.graph_time[0] < cut:
            dp.graph_time.pop(0)
            dp.graph_temp_supply.pop(0)
            dp.graph_temp_return.pop(0)
            dp.graph_press_supply.pop(0)

        # обновляем кривые
        dp.curve_temp_supply.setData(dp.graph_time, dp.graph_temp_supply)
        dp.curve_temp_return.setData(dp.graph_time, dp.graph_temp_return)
        dp.curve_press_supply.setData(dp.graph_time, dp.graph_press_supply)

        # автоскролл (если юзер не двигает график)
        if not dp.user_is_panning:
            dp.main_plot.setXRange(cut, t, padding=0)

    def load_initial_history(self):
        dp = self.dashboard_page

        # если уже загружено для этого котла — не грузим повторно
        if self._history_loaded and self._history_loaded_for == dp.selected_boiler:
            return

        hours = int(getattr(dp, "graph_window_hours", 4))
        now = datetime.now()

        # какой котёл выбран
        n = int(dp.selected_boiler)

        # какие PID нам нужны
        pid_supply = self.param_ids_by_key.get(f"boiler_{n}")
        pid_return = self.param_ids_by_key.get(f"boiler_{n}_return")
        pid_press = self.param_ids_by_key.get(f"boiler_{n}_pressure")

        ids = [pid for pid in (pid_supply, pid_return, pid_press) if pid]

        if not ids:
            print("[HISTORY] нет настроенных PID для графика")
            return

        # режим периода или live
        if getattr(dp, "period_mode", False) and dp.period_start_ts and dp.period_end_ts:
            start_ts = int(dp.period_start_ts)
            end_ts = int(dp.period_end_ts)
        else:
            start = now - timedelta(hours=hours)
            start_ts = int(start.timestamp())
            end_ts = int(now.timestamp())

        try:
            hist = self.api_client.get_history(
                ids,
                start_ts,
                end_ts,
                step_sec=60,
            )
        except Exception as e:
            print("[HISTORY] load failed:", e)
            return

        # получаем серии
        series_supply = hist.get(pid_supply, []) if pid_supply else []
        series_return = hist.get(pid_return, []) if pid_return else []
        series_press = hist.get(pid_press, []) if pid_press else []

        print("[HISTORY] lens:", len(series_supply), len(series_return), len(series_press))

        if not series_supply and not series_return and not series_press:
            print("[HISTORY] пустые серии")
            return

        # очищаем
        dp.graph_time.clear()
        dp.graph_temp_supply.clear()
        dp.graph_temp_return.clear()
        dp.graph_press_supply.clear()

        # превращаем в словари
        d_supply = dict(series_supply)
        d_return = dict(series_return)
        d_press = dict(series_press)

        all_ts = sorted(set(d_supply) | set(d_return) | set(d_press))

        last_sup = last_ret = last_pre = 0.0

        for ts in all_ts:
            if ts in d_supply:
                last_sup = d_supply[ts]
            if ts in d_return:
                last_ret = d_return[ts]
            if ts in d_press:
                last_pre = d_press[ts]

            dp.graph_time.append(ts)
            dp.graph_temp_supply.append(last_sup)
            dp.graph_temp_return.append(last_ret)
            dp.graph_press_supply.append(last_pre)

        dp.curve_temp_supply.setData(dp.graph_time, dp.graph_temp_supply)
        dp.curve_temp_return.setData(dp.graph_time, dp.graph_temp_return)
        dp.curve_press_supply.setData(dp.graph_time, dp.graph_press_supply)

        dp.main_plot.setXRange(start_ts, end_ts, padding=0)
        dp.user_is_panning = bool(getattr(dp, "period_mode", False))

        self._history_loaded = True
        self._history_loaded_for = n